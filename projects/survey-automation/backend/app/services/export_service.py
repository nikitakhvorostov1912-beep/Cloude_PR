"""Сервис экспорта документов."""

from __future__ import annotations

import aiofiles
import json
import logging
import shutil
import tempfile
from pathlib import Path
from typing import Any

from app.config import ProjectDir, get_project_dir
from app.docs.process_doc import ProcessDocGenerator
from app.exceptions import ExportError, NotFoundError, ProcessingError

logger = logging.getLogger(__name__)


class ExportService:
    """Экспорт данных проекта в различные форматы документов.

    Поддерживает генерацию Word, Excel, Visio и ZIP-архивов
    с результатами анализа бизнес-процессов.
    """

    # ------------------------------------------------------------------
    # Экспорт Visio (.vsdx)
    # ------------------------------------------------------------------

    async def export_visio(
        self,
        project_id: str,
        process_id: str,
    ) -> Path:
        """Генерирует Visio-файл (.vsdx) для одного процесса.

        Args:
            project_id: Идентификатор проекта.
            process_id: Идентификатор процесса.

        Returns:
            Путь к сгенерированному .vsdx файлу.

        Raises:
            NotFoundError: Если процесс не найден.
            ExportError: При ошибке генерации Visio.
        """
        project_dir = self._ensure_project_exists(project_id)

        # Проверяем наличие BPMN JSON или файла процесса
        bpmn_json = await self._load_process_bpmn(project_dir, process_id)

        project_dir.ensure_dirs()
        output_path = project_dir.get_visio_path(process_id)

        try:
            from app.visio import generate_visio

            generate_visio(bpmn_json, output_path)
        except Exception as exc:
            raise ExportError(
                f"Ошибка генерации Visio для процесса: {process_id}",
                detail=str(exc),
            ) from exc

        logger.info("Visio сгенерирован: %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Экспорт описания процессов (Word)
    # ------------------------------------------------------------------

    async def export_process_doc(self, project_id: str) -> Path:
        """Генерирует Word-документ с описанием всех процессов.

        Args:
            project_id: Идентификатор проекта.

        Returns:
            Путь к сгенерированному .docx файлу.

        Raises:
            NotFoundError: Если процессы не найдены.
            ExportError: При ошибке генерации документа.
        """
        project_dir = self._ensure_project_exists(project_id)
        processes = await self._load_all_processes(project_dir)

        if not processes:
            raise NotFoundError(
                "Бизнес-процессы не найдены. Сначала выполните извлечение процессов.",
                detail={"project_id": project_id},
            )

        project_data = await self._load_project_meta(project_dir)
        project_name = project_data.get("name", "Проект")

        project_dir.ensure_dirs()
        output_path = project_dir.get_output_path("описание_процессов", ext=".docx")

        try:
            generator = ProcessDocGenerator()
            generator.generate(
                processes=processes,
                output_path=output_path,
                project_name=project_name,
            )
        except Exception as exc:
            raise ExportError(
                "Ошибка генерации документа описания процессов",
                detail=str(exc),
            ) from exc

        logger.info("Документ процессов сгенерирован: %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Экспорт требований (Excel)
    # ------------------------------------------------------------------

    async def export_requirements_excel(self, project_id: str) -> Path:
        """Генерирует Excel-файл с листом требований.

        Args:
            project_id: Идентификатор проекта.

        Returns:
            Путь к сгенерированному .xlsx файлу.

        Raises:
            NotFoundError: Если требования не найдены.
            ExportError: При ошибке генерации.
        """
        project_dir = self._ensure_project_exists(project_id)
        requirements_data = await self._load_requirements(project_dir)

        if not requirements_data:
            raise NotFoundError(
                "Требования не найдены. Сначала выполните генерацию требований.",
                detail={"project_id": project_id},
            )

        project_dir.ensure_dirs()
        output_path = project_dir.get_output_path("требования", ext=".xlsx")

        try:
            from app.docs.doc_generator import (
                EXCEL_ALIGNMENT_LEFT,
                EXCEL_BORDER,
                EXCEL_FILLS,
                EXCEL_FONTS,
                DocGenerator,
            )
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Требования"

            headers = [
                "ID", "Тип", "Название", "Описание", "Категория",
                "Приоритет", "Источник", "Критерии приёмки",
                "Трудоёмкость (ч)", "Подсистема 1С",
            ]
            DocGenerator._write_excel_header(ws, headers)

            for row_idx, req in enumerate(requirements_data, start=2):
                ws.cell(row=row_idx, column=1, value=req.get("id", ""))
                ws.cell(row=row_idx, column=2, value=req.get("type", ""))
                ws.cell(row=row_idx, column=3, value=req.get("name", ""))
                ws.cell(row=row_idx, column=4, value=req.get("description", ""))
                ws.cell(row=row_idx, column=5, value=req.get("category", ""))
                ws.cell(row=row_idx, column=6, value=req.get("priority", ""))
                ws.cell(row=row_idx, column=7, value=req.get("source", ""))

                criteria = req.get("acceptance_criteria", [])
                if isinstance(criteria, list):
                    criteria = "\n".join(f"- {c}" for c in criteria)
                ws.cell(row=row_idx, column=8, value=str(criteria))

                ws.cell(row=row_idx, column=9, value=req.get("effort_hours", ""))
                ws.cell(row=row_idx, column=10, value=req.get("erp_subsystem", ""))

                # Цвет приоритета
                priority = str(req.get("priority", "")).lower()
                priority_key = priority if priority in EXCEL_FILLS else None
                if priority_key:
                    ws.cell(row=row_idx, column=6).fill = EXCEL_FILLS[priority_key]

                # Рамки и выравнивание
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = EXCEL_BORDER
                    cell.alignment = EXCEL_ALIGNMENT_LEFT
                    cell.font = EXCEL_FONTS["normal"]

            DocGenerator._auto_adjust_columns(ws)
            DocGenerator._freeze_panes(ws)
            DocGenerator._add_autofilter(ws)

            wb.save(str(output_path))

        except Exception as exc:
            raise ExportError(
                "Ошибка генерации Excel с требованиями",
                detail=str(exc),
            ) from exc

        logger.info("Excel требований сгенерирован: %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Экспорт требований (Word)
    # ------------------------------------------------------------------

    async def export_requirements_word(self, project_id: str) -> Path:
        """Генерирует Word-документ с требованиями.

        Args:
            project_id: Идентификатор проекта.

        Returns:
            Путь к сгенерированному .docx файлу.

        Raises:
            NotFoundError: Если требования не найдены.
            ExportError: При ошибке генерации.
        """
        project_dir = self._ensure_project_exists(project_id)
        requirements_data = await self._load_requirements(project_dir)

        if not requirements_data:
            raise NotFoundError(
                "Требования не найдены. Сначала выполните генерацию требований.",
                detail={"project_id": project_id},
            )

        project_data = await self._load_project_meta(project_dir)
        project_name = project_data.get("name", "Проект")

        project_dir.ensure_dirs()
        output_path = project_dir.get_output_path("лист_требований", ext=".docx")

        try:
            from app.docs.doc_generator import safe_str

            # Используем базовый генератор для утилит
            gen = ProcessDocGenerator()
            doc = gen._create_document(title=f"Лист требований — {project_name}")
            gen._add_title_page(doc, "Лист требований", project_name)
            gen._add_table_of_contents(doc)
            gen._add_header(doc, f"{project_name} — Лист требований")
            gen._add_footer_with_pages(doc)

            # Группируем требования по типу
            by_type: dict[str, list[dict[str, Any]]] = {}
            for req in requirements_data:
                rtype = req.get("type", "Прочие")
                by_type.setdefault(rtype, []).append(req)

            type_names = {
                "FR": "Функциональные требования",
                "NFR": "Нефункциональные требования",
                "IR": "Интеграционные требования",
            }

            for rtype, reqs in by_type.items():
                section_name = type_names.get(rtype, rtype)
                doc.add_heading(section_name, level=1)

                for req in reqs:
                    req_id = req.get("id", "")
                    req_name = req.get("name", "")
                    doc.add_heading(f"{req_id}: {req_name}", level=2)

                    # Описание
                    desc = req.get("description", "")
                    if desc:
                        doc.add_paragraph(desc)

                    # Основные поля
                    fields = [
                        ("Приоритет", req.get("priority", "")),
                        ("Категория", req.get("category", "")),
                        ("Источник", req.get("source", "")),
                        ("Трудоёмкость", f"{req.get('effort_hours', '—')} ч"),
                        ("Подсистема 1С", req.get("erp_subsystem", "")),
                    ]

                    table = doc.add_table(rows=len(fields), cols=2)
                    table.style = "Table Grid"
                    for row, (label, value) in zip(table.rows, fields):
                        row.cells[0].text = label
                        row.cells[1].text = safe_str(value)
                    gen._format_table_header(table)

                    # Критерии приёмки
                    criteria = req.get("acceptance_criteria", [])
                    if criteria:
                        doc.add_heading("Критерии приёмки", level=3)
                        for criterion in criteria:
                            doc.add_paragraph(str(criterion), style="List Bullet")

            doc.save(str(output_path))

        except Exception as exc:
            raise ExportError(
                "Ошибка генерации Word с требованиями",
                detail=str(exc),
            ) from exc

        logger.info("Word требований сгенерирован: %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Экспорт GAP-отчёта (Excel)
    # ------------------------------------------------------------------

    async def export_gap_report(self, project_id: str) -> Path:
        """Генерирует Excel-отчёт по GAP-анализу.

        Args:
            project_id: Идентификатор проекта.

        Returns:
            Путь к сгенерированному .xlsx файлу.

        Raises:
            NotFoundError: Если результаты GAP-анализа не найдены.
            ExportError: При ошибке генерации.
        """
        project_dir = self._ensure_project_exists(project_id)
        gaps = await self._load_gaps(project_dir)

        if not gaps:
            raise NotFoundError(
                "Результаты GAP-анализа не найдены. Сначала выполните GAP-анализ.",
                detail={"project_id": project_id},
            )

        project_dir.ensure_dirs()
        output_path = project_dir.get_output_path("gap_анализ", ext=".xlsx")

        try:
            from app.docs.doc_generator import (
                EXCEL_ALIGNMENT_LEFT,
                EXCEL_BORDER,
                EXCEL_FILLS,
                EXCEL_FONTS,
                DocGenerator,
            )
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "GAP-анализ"

            headers = [
                "Процесс", "Шаг №", "Шаг", "Покрытие",
                "Механизм 1С", "Разрыв (GAP)", "Рекомендация",
                "Трудоёмкость", "Приоритет", "Риски",
            ]
            DocGenerator._write_excel_header(ws, headers)

            row_idx = 2
            for gap in gaps:
                process_name = gap.get("process_name", gap.get("process_id", ""))
                step_analysis = gap.get("step_analysis", [])

                for step in step_analysis:
                    ws.cell(row=row_idx, column=1, value=process_name)
                    ws.cell(row=row_idx, column=2, value=step.get("step_order", ""))
                    ws.cell(row=row_idx, column=3, value=step.get("step_name", ""))
                    ws.cell(row=row_idx, column=4, value=step.get("coverage", ""))
                    ws.cell(row=row_idx, column=5, value=step.get("erp_mechanism", ""))
                    ws.cell(row=row_idx, column=6, value=step.get("gap_description", ""))
                    ws.cell(row=row_idx, column=7, value=step.get("recommendation", ""))
                    ws.cell(row=row_idx, column=8, value=step.get("effort", ""))
                    ws.cell(row=row_idx, column=9, value=step.get("priority", ""))

                    risks = step.get("risks", [])
                    if isinstance(risks, list):
                        risks = "; ".join(risks)
                    ws.cell(row=row_idx, column=10, value=str(risks))

                    # Цвет покрытия
                    coverage = str(step.get("coverage", "")).lower()
                    coverage_colors = {
                        "full": "green",
                        "partial": "yellow",
                        "custom": "should",
                        "absent": "red",
                    }
                    fill_key = coverage_colors.get(coverage)
                    if fill_key and fill_key in EXCEL_FILLS:
                        ws.cell(row=row_idx, column=4).fill = EXCEL_FILLS[fill_key]

                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col)
                        cell.border = EXCEL_BORDER
                        cell.alignment = EXCEL_ALIGNMENT_LEFT
                        cell.font = EXCEL_FONTS["normal"]

                    row_idx += 1

            DocGenerator._auto_adjust_columns(ws)
            DocGenerator._freeze_panes(ws)
            DocGenerator._add_autofilter(ws)

            wb.save(str(output_path))

        except Exception as exc:
            raise ExportError(
                "Ошибка генерации Excel GAP-отчёта",
                detail=str(exc),
            ) from exc

        logger.info("Excel GAP-отчёт сгенерирован: %s", output_path)
        return output_path

    # ------------------------------------------------------------------
    # Экспорт всего (ZIP)
    # ------------------------------------------------------------------

    async def export_all(self, project_id: str) -> Path:
        """Генерирует ZIP-архив со всеми документами проекта.

        Включает: описание процессов (Word), требования (Excel + Word),
        GAP-отчёт (Excel), BPMN-файлы, SVG-диаграммы.

        Args:
            project_id: Идентификатор проекта.

        Returns:
            Путь к сгенерированному .zip файлу.

        Raises:
            ExportError: При ошибке создания архива.
        """
        project_dir = self._ensure_project_exists(project_id)
        project_dir.ensure_dirs()

        # Генерируем все доступные документы, игнорируя ошибки
        generated_files: list[Path] = []

        # Описание процессов
        try:
            path = await self.export_process_doc(project_id)
            generated_files.append(path)
        except (NotFoundError, ExportError) as exc:
            logger.warning("Пропущен экспорт процессов: %s", exc.message)

        # Требования Excel
        try:
            path = await self.export_requirements_excel(project_id)
            generated_files.append(path)
        except (NotFoundError, ExportError) as exc:
            logger.warning("Пропущен экспорт требований (Excel): %s", exc.message)

        # Требования Word
        try:
            path = await self.export_requirements_word(project_id)
            generated_files.append(path)
        except (NotFoundError, ExportError) as exc:
            logger.warning("Пропущен экспорт требований (Word): %s", exc.message)

        # GAP-отчёт
        try:
            path = await self.export_gap_report(project_id)
            generated_files.append(path)
        except (NotFoundError, ExportError) as exc:
            logger.warning("Пропущен экспорт GAP-отчёта: %s", exc.message)

        # Собираем BPMN и SVG файлы
        bpmn_files: list[Path] = []
        if project_dir.bpmn.is_dir():
            bpmn_files = sorted(
                p for p in project_dir.bpmn.iterdir()
                if p.is_file() and p.suffix.lower() in {".bpmn", ".svg"}
            )

        if not generated_files and not bpmn_files:
            raise ExportError(
                "Нет данных для экспорта. Выполните анализ перед экспортом.",
                detail={"project_id": project_id},
            )

        # Создаём ZIP-архив
        zip_path = project_dir.get_output_path("полный_экспорт", ext=".zip")

        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                tmp_path = Path(tmp_dir) / "export"
                tmp_path.mkdir()

                # Копируем сгенерированные документы
                docs_dir = tmp_path / "documents"
                docs_dir.mkdir()
                for fp in generated_files:
                    if fp.is_file():
                        shutil.copy2(fp, docs_dir / fp.name)

                # Копируем BPMN файлы
                if bpmn_files:
                    bpmn_dir = tmp_path / "bpmn"
                    bpmn_dir.mkdir()
                    for fp in bpmn_files:
                        shutil.copy2(fp, bpmn_dir / fp.name)

                # Создаём архив
                archive_base = str(zip_path).removesuffix(".zip")
                shutil.make_archive(archive_base, "zip", tmp_path)

        except Exception as exc:
            raise ExportError(
                "Ошибка создания ZIP-архива",
                detail=str(exc),
            ) from exc

        logger.info(
            "Полный экспорт создан: %s (%d документов, %d BPMN файлов)",
            zip_path,
            len(generated_files),
            len(bpmn_files),
        )
        return zip_path

    # ------------------------------------------------------------------
    # Приватные методы
    # ------------------------------------------------------------------

    @staticmethod
    def _ensure_project_exists(project_id: str) -> ProjectDir:
        """Проверяет существование проекта."""
        project_dir = get_project_dir(project_id)
        if not project_dir.exists():
            raise NotFoundError(
                f"Проект не найден: {project_id}",
                detail={"project_id": project_id},
            )
        return project_dir

    async def _load_project_meta(self, project_dir: ProjectDir) -> dict[str, Any]:
        """Загружает метаданные проекта из project.json."""
        meta_path = project_dir.root / "project.json"
        if not meta_path.is_file():
            return {}
        try:
            async with aiofiles.open(meta_path, "r", encoding="utf-8") as f:
                content = await f.read()
            return json.loads(content)
        except Exception:
            return {}

    async def _load_all_processes(self, project_dir: ProjectDir) -> list[dict[str, Any]]:
        """Загружает все процессы проекта."""
        all_path = project_dir.processes / "_all_processes.json"
        if all_path.is_file():
            try:
                async with aiofiles.open(all_path, "r", encoding="utf-8") as f:
                    content = await f.read()
                data = json.loads(content)
                if isinstance(data, list):
                    return data
            except Exception as exc:
                logger.warning("Не удалось загрузить %s: %s", all_path.name, exc)

        if not project_dir.processes.is_dir():
            return []

        processes: list[dict[str, Any]] = []
        for jf in sorted(project_dir.processes.iterdir()):
            if (
                jf.is_file()
                and jf.suffix.lower() == ".json"
                and not jf.name.startswith("_")
                and not jf.name.endswith("_gap.json")
                and not jf.name.endswith("_tobe.json")
                and not jf.name.endswith("_bpmn.json")
            ):
                try:
                    async with aiofiles.open(jf, "r", encoding="utf-8") as f:
                        content = await f.read()
                    data = json.loads(content)
                    if isinstance(data, dict):
                        processes.append(data)
                except Exception as exc:
                    logger.warning("Не удалось прочитать процесс %s: %s", jf.name, exc)
                    continue
        return processes

    async def _load_requirements(self, project_dir: ProjectDir) -> list[dict[str, Any]]:
        """Загружает лист требований."""
        req_path = project_dir.processes / "_requirements.json"
        if not req_path.is_file():
            return []
        try:
            async with aiofiles.open(req_path, "r", encoding="utf-8") as f:
                content = await f.read()
            data = json.loads(content)
            if isinstance(data, dict):
                return data.get("requirements", [data])
            if isinstance(data, list):
                return data
        except Exception as exc:
            logger.warning("Не удалось загрузить требования %s: %s", req_path.name, exc)
        return []

    async def _load_gaps(self, project_dir: ProjectDir) -> list[dict[str, Any]]:
        """Загружает результаты GAP-анализа."""
        gaps_path = project_dir.processes / "_gap_analysis.json"
        if not gaps_path.is_file():
            return []
        try:
            async with aiofiles.open(gaps_path, "r", encoding="utf-8") as f:
                content = await f.read()
            data = json.loads(content)
            if isinstance(data, list):
                return data
        except Exception as exc:
            logger.warning("Не удалось загрузить GAP-анализ %s: %s", gaps_path.name, exc)
        return []

    async def _load_process_bpmn(
        self,
        project_dir: ProjectDir,
        process_id: str,
    ) -> dict[str, Any]:
        """Загружает BPMN JSON для одного процесса."""
        # Сначала ищем _bpmn.json
        bpmn_path = project_dir.processes / f"{process_id}_bpmn.json"
        if not bpmn_path.is_file():
            # Пробуем обычный JSON процесса
            bpmn_path = project_dir.get_process_path(process_id)

        if not bpmn_path.is_file():
            raise NotFoundError(
                f"Данные процесса не найдены: {process_id}",
                detail={"project_id": str(project_dir.root.name), "process_id": process_id},
            )

        try:
            async with aiofiles.open(bpmn_path, "r", encoding="utf-8") as f:
                content = await f.read()
            data = json.loads(content)
        except Exception as exc:
            raise ProcessingError(
                f"Ошибка чтения данных процесса: {process_id}",
                detail=str(exc),
            ) from exc

        if not isinstance(data, dict):
            raise ProcessingError(
                f"Некорректные данные процесса: {process_id}",
            )
        return data
